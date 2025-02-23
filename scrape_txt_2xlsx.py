import os
import copy
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from datetime import datetime


class Txt_Scraper:
    def __init__(self, folder):
        self.folder = folder
        self.all_data = []
        self.count = 0
        self.template = {
            "File": "",
            "Date": "",
            "Invoice Number": "",
            "Amount": "",
            "Currency": "",
            "Country": "",
        }
        self.get_all_txt_file()
        self.report()
        self.save_as_xlsx()

    @staticmethod
    def parse_date(date_str, currency):
        MM_DD_YYYY_CURRENCIES = {"USD"}  # Mainly the US
        # Define currencies that follow DD-MM-YYYY
        DD_MM_YYYY_CURRENCIES = {"EUR", "GBP", "AUD", "CAD"}
        try:
            # Format with full month name (e.g., 03-Aug-2024)
            if "-" in date_str and any(c.isalpha() for c in date_str):
                return datetime.strptime(date_str, "%d-%b-%Y").date()

            # ISO format (e.g., 2024-05-02)
            if "-" in date_str and date_str.count("-") == 2 and date_str[:4].isdigit():
                return datetime.strptime(date_str, "%Y-%m-%d").date()

            # Ambiguous format (e.g., 02-01-2023)
            if currency in MM_DD_YYYY_CURRENCIES:
                return datetime.strptime(date_str, "%m-%d-%Y").date()  # MM-DD-YYYY
            elif currency in DD_MM_YYYY_CURRENCIES:
                return datetime.strptime(date_str, "%d-%m-%Y").date()  # DD-MM-YYYY

            # Default fallback (try MM-DD-YYYY first, then DD-MM-YYYY)
            try:
                return datetime.strptime(date_str, "%m-%d-%Y").date()
            except ValueError:
                return datetime.strptime(date_str, "%d-%m-%Y").date()

        except ValueError:
            print(f"Error parsing date: {date_str} (Currency: {currency})")
            return None

    def get_all_txt_file(self):
        if not os.path.isdir(self.folder):
            print(f"Error: The folder '{self.folder}' does not exist.")
            return

        for filename in os.listdir(self.folder):
            if filename.lower().endswith(".txt"):
                file_path = os.path.join(self.folder, filename)
                print(Path(re.sub(".txt", ".pdf", file_path)).as_uri())
                print(Path(file_path).as_uri())
                self.read_txt_file(file_path, re.sub(".txt", ".pdf", filename))
                self.count += 1

    def read_txt_file(self, txt_file, filename):
        data = copy.deepcopy(self.template)

        f = open(txt_file, "r")
        lines = f.readlines()
        i = 0
        for line in lines:
            # print(line.strip())
            # m1 = re.search(r"Invoice Number:\s*([A-Z0-9]{5,}-[0-9]+)", line)
            # m2 = re.search(r"Invoice Date: (\d+-\d+-\d+)\s*$", line)
            # m3 = re.search(r"Invoice Currency: ([A-Z]{3})\s*$", line)

            m1 = re.search(r":\s*([A-Z0-9]{5,}-[0-9]+)", line)
            m2 = re.search(r":\s*(\d+-(\d+|[A-Za-z]+)-\d+)\s*$", line)
            m3 = re.search(r":\s*([A-Z]{3})\s*$", line)
            m4 = re.search(r"((\d+[,.])*\d+[,.]\d+)", line)
            data["File"] = filename
            if m1 is not None:
                data["Invoice Number"] = m1.group(1)
            if m2 is not None:
                data["Date"] = m2.group(1)
            if m3 is not None:
                data["Currency"] = m3.group(1)
                data["Country"] = self.currency_to_country(m3.group(1))
            if i <= 1 and m4 is not None:
                data["Amount"] = m4.group(1)
            elif i <= 1 and data["Amount"] == "":
                print(f"  -->{line}")
            i += 1

        if data["Amount"] == "":
            print("-->>Amount missed")

        f.close()
        data["Date"] = self.parse_date(data["Date"],data["Currency"])

        self.all_data.append(data)
        print(data)

    @staticmethod
    def currency_to_country(currency):
        data = {
            "USD": "US",
            "JPY": "JAPAN",
        }
        if currency in data:
            return data[currency]
        else:
            return ""

    def report(self):
        print(f"\nNumber of file: {self.count}")

    def save_as_xlsx(self):
        book = Workbook()
        sheet = book.active

        # ------------------Insert data -------------------------
        sheet.append(list(self.template.keys()))

        for data in self.all_data:
            sheet.append(list(data.values()))

        # -------------------- Layout -------------------------
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        column_width = {"A": 60, "B": 25, "C": 20, "D": 30, "E": 20, "F": 20}
        for row in range(sheet.max_row):
            for col in range(len(column_width)):
                sheet.cell(row + 1, col + 1).border = thin_border

        for col in column_width:
            sheet.column_dimensions[col].width = column_width[col]
            sheet[f"{col}1"].font = Font(bold=True)
            sheet[f"{col}1"].fill = PatternFill(start_color="00A9E6", fill_type="solid")
            sheet[f"{col}1"].border = thin_border

        # ------------------------- Save ----------------------
        book.save(r"C:\Users\NJAKA\Desktop\INVOICE.xlsx")
        book.close()



# Txt_Scraper(r"C:\Users\NJAKA\Desktop\invoice")
Txt_Scraper(r"C:\Users\NJAKA\Downloads\Compressed\Advertising Invoices")


