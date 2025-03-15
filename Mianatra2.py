import wx
import os
import load_exo
import random
import vlc
import re
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import io
from PIL import Image
import subprocess
import json


class MediaPlayer:
    def __init__(self, panel, parent_frame):
        self.Instance = vlc.Instance()
        self.player = self.Instance.media_player_new()
        self.is_playing = False
        self.panel = panel
        self.file_path = ""
        self.parent_frame = parent_frame
        self.media_player_events = self.player.event_manager()
        self.media_player_events.event_attach(
            vlc.EventType.MediaPlayerEndReached, self.on_media_end
        )

    def play_media(self, file_path):
        self.file_path = file_path
        if self.is_playing:
            self.player.stop()
            self.is_playing = False

        Media = self.Instance.media_new(file_path)
        self.player.set_media(Media)

        # Get the handle of the panel
        handle = self.panel.GetHandle()
        self.player.set_hwnd(handle)

        # Set video dimensions
        # self.player.video_set_aspect_ratio("854:x")

        self.player.play()
        self.is_playing = True

        if re.search("mp3$", file_path) is not None:
            self.panel.Hide()
        else:
            self.panel.Show()
        self.parent_frame.playing_video = True

    def on_media_end(self, event):
        if self.is_playing:  # Check if still playing to avoid multiple calls
            self.is_playing = False
            self.panel.Hide()  # Hide the panel when video ends
            print("Video playback finished. Panel hidden.")
            if re.search("course.*\\.(mp4|avi)$", self.file_path, re.I):
                self.parent_frame.after_video_ends()
                self.parent_frame.playing_video = False


class MyFrame(wx.Frame):
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title=title, size=(800, 450), style=wx.DEFAULT_FRAME_STYLE | wx.MAXIMIZE)
        # Change the current working directory
        os.chdir(r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2")
        # os.chdir(r"C:\Users\NJAKA\Mianatra2")
        self.app_data_folder = f"{os.getenv('LOCALAPPDATA')}/Mianatra2"
        self.settings_file = f"{self.app_data_folder}/Minatra2_setting.json"

        self.bitmap_buttons = []
        self.playing_video = False
        self.ID_SETTING = wx.NewIdRef()
        self.ID_RESTART = wx.NewIdRef()

        self.choice_answer_available = False
        self.log_txt_file = "log_exo_itokiana.txt"
        self.log_excel_file = "log_exo_itokiana.xlsx"

        # Load background image
        img_path = os.path.join("images", "orange_ice_mint.jpg")
        if os.path.isfile(img_path):
            self.background_image = wx.Image(img_path, wx.BITMAP_TYPE_ANY)
        else:
            self.background_image = wx.Image(854, 480)

        # Limit the width of the background image to 800 pixels while maintaining the aspect ratio
        if self.background_image.GetWidth() > 854:
            new_width = 854
            new_height = int(854 * self.background_image.GetHeight() / self.background_image.GetWidth())
            self.background_image = self.background_image.Scale(new_width, new_height, wx.IMAGE_QUALITY_HIGH)

        self.background_bitmap = wx.Bitmap(self.background_image)

        # Create home panel
        self.home_panel = wx.Panel(self, -1)
        self.dc = wx.ClientDC(self.home_panel)

        # Create a menu bar
        menu_bar = wx.MenuBar()

        # Create a file menu
        file_menu = wx.Menu()
        restart_item = file_menu.Append(self.ID_RESTART, "Restart\tCtrl+R", "Open settings")
        restart_item = file_menu.Append(wx.ID_ANY, "Dashboard", "Open settings")
        restart_item = file_menu.Append(wx.ID_ANY, "Test Exo", "Open settings")
        file_menu.AppendSeparator()
        file_menu.Append(self.ID_SETTING, "Settings\tCtrl+Shift+S", "Open settings")
        file_menu.AppendSeparator()
        file_menu.Append(wx.ID_EXIT, "Exit\tAlt+F4", "Exit the application")
        menu_bar.Append(file_menu, "&File")

        # Create a help menu
        help_menu = wx.Menu()
        help_menu.Append(wx.ID_ABOUT, "About\tF1", "About this application")
        menu_bar.Append(help_menu, "&Help")

        # Set the menu bar
        self.SetMenuBar(menu_bar)

        # Create a toolbar
        toolbar = self.CreateToolBar()

        # Add some tools to the toolbar
        tool1 = toolbar.AddTool(wx.ID_ANY, "Tool 1", wx.ArtProvider.GetBitmap(wx.ART_WX_LOGO, wx.ART_BUTTON), "Tool 1 tooltip")
        tool2 = toolbar.AddTool(wx.ID_ANY, "Tool 2", wx.ArtProvider.GetBitmap(wx.ART_FOLDER, wx.ART_BUTTON), "Tool 2 tooltip")

        # Realize the toolbar
        toolbar.Realize()
        toolbar.SetBackgroundColour(wx.Colour(184, 197, 239))
        self.SetToolBar(toolbar)

        # Create a status bar
        self.status_bar = self.CreateStatusBar()
        self.status_bar.SetFieldsCount(5)  # Split into 4 parts
        self.status_bar.SetStatusWidths([-1, -1, -1, -1, 100])  # Set relative widths

        # Add a progress bar to the last part
        self.add_progress_bar_to_status_bar()
        self.progress_bar.SetValue(0)

        # Create a sizer to layout controls
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Add a StaticBitmap for the background
        self.background_staticbitmap = wx.StaticBitmap(self.home_panel, -1, self.background_bitmap)
        main_sizer.Add(self.background_staticbitmap, 1, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 10)  # Set proportion to 1 to make it expand

        self.bitmap_sizer = wx.BoxSizer(wx.HORIZONTAL)
        main_sizer.Add(self.bitmap_sizer, 0, wx.ALIGN_CENTER | wx.BOTTOM, 20)

        self.video_panel = wx.Panel(self.home_panel, style=wx.SIMPLE_BORDER)
        # self.video_panel.SetMaxSize((854, 480))
        main_sizer.Add(self.video_panel, 1, wx.EXPAND | wx.ALL, 0)

        self.video_panel.Hide()
        # Create textCtrl
        self.valiny = wx.TextCtrl(self.home_panel, -1, "", size=(854, -1), style=wx.TE_LEFT | wx.TE_PROCESS_ENTER)
        self.valiny.SetFont(wx.Font(21, wx.DEFAULT, wx.NORMAL, wx.NORMAL))
        self.valiny.SetFocus()  # Set initial focus to textCtrl
        main_sizer.Add(self.valiny, 0, wx.ALIGN_CENTER | wx.BOTTOM, 10)

        # Create OK button
        self.ok_button = wx.Button(self.home_panel, -1, "START")
        self.ok_button.SetCursor(wx.Cursor(wx.CURSOR_HAND))
        font = wx.Font(20, wx.DEFAULT, wx.NORMAL, wx.NORMAL)
        self.ok_button.SetFont(font)
        main_sizer.Add(self.ok_button, 0, wx.ALIGN_CENTER | wx.BOTTOM, 20)

        self.player = MediaPlayer(self.video_panel, self)
        # self.video_panel.SetBackgroundColour(wx.Colour(200, 0, 0))

        self.home_panel.SetSizer(main_sizer)
        self.valiny.Hide()
        self.create_bitmap_buttons()

        self.home_panel.Layout()

        self.Bind(wx.EVT_TOOL, self.OnTool1, tool1)
        self.Bind(wx.EVT_TOOL, self.OnTool2, tool2)
        self.Bind(wx.EVT_SIZE, self.on_resize)
        self.valiny.Bind(wx.EVT_TEXT_ENTER, self.on_enter_pressed)
        self.ok_button.Bind(wx.EVT_BUTTON, self.on_ok_button)
        self.background_staticbitmap.Bind(wx.EVT_LEFT_DOWN, self.on_background_click)
        self.Bind(wx.EVT_MENU, self.OnRestart, id=self.ID_RESTART)
        self.Bind(wx.EVT_MENU, self.OnSettings, id=self.ID_SETTING)
        self.Bind(wx.EVT_MENU, self.OnExit, id=wx.ID_EXIT)
        self.Bind(wx.EVT_MENU, self.OnAbout, id=wx.ID_ABOUT)
        self.home_panel.Bind(wx.EVT_MOTION, self.on_home_panel_motion)

        self.all_exo = {}
        self.current_exo_name = ""
        self.exo_done = []
        self.exo_list = []
        self.all_exo_completed = False

        self.stage_min = 2
        self.stage_max = 5
        self.stage_rand = False
        self.stage_level = 0
        self.stage_type = "entry"
        self.stage_case_sensitive = False
        self.stage_library = None
        self.stage_comment = "Comment"

        self.stage_current_index = None
        self.stage_index_done = []

        # self.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        # self.SetBackgroundColour(wx.Colour(200, 200, 200))
        self.Show()
        self.Maximize()
        # self.load_settings()
        self.ok_button.SetFocus()

    def load_settings(self):
        print("---->>load_settings")
        if os.path.isfile(self.settings_file):
            with open(self.settings_file) as json_file:
                data = json.load(json_file)
                if 'exo_folder' in data and os.path.isdir(f"{data['exo_folder']}"):
                    os.chdir(data['exo_folder'])
        else:
            self.save_settings()

    def save_settings(self, exo_dir=os.getcwd()):
        print("---->>save_settings")

        try:
            os.mkdir(self.app_data_folder)
        except Exception as e:
            print(f"An error was occurred (mkdir): {e}")

        try:
            if os.path.isdir(exo_dir):
                data = {'exo_folder': exo_dir}

                with open(self.settings_file, 'w') as outfile:
                    json.dump(data, outfile, indent=3)
        except Exception as e:
            wx.MessageBox(f"Cannot create the settings file:\n'{self.settings_file}' --> {e}")

    def after_video_ends(self):  # Method to be called after video ends
        print("->Next Step in MyFrame")
        self.valiny.SetValue("")
        self.playing_video = False
        if self.current_exo_name is not None:
            self.get_exo_index()
            if self.stage_current_index is not None:
                self.display_exo()

    def update_log_file(self, text):
        text_list = [text]
        if type(text) is list:
            text_list = text
            text = " | ".join(map(str, text))
        daty = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        try:
            with open(self.log_txt_file, "a") as file:
                file.write(f"{daty} | {text}" + "\n")
        except PermissionError:
            pass

        # Log for excel
        print(f"-->Excel: {text_list}")
        if os.path.isfile(self.log_excel_file):
            book = load_workbook(self.log_excel_file)
            sheet = book.active
        else:
            book = Workbook()
            sheet = book.active
            sheet.append(("Daty", "Marina/Diso", "Fanazarana", "Index", "Fanazavana", "Valiny natao"))
            column_width = {"A": 25, "B": 15, "C": 30, "D": 15, "E": 30, "F": 30}
            for col in column_width:
                sheet.column_dimensions[col].width = column_width[col]
                sheet[f"{col}1"].font = Font(bold=True)
                sheet[f"{col}1"].fill = PatternFill(start_color="00A9E6", fill_type="solid")
        sheet.append([daty] + text_list)
        try:
            book.save(self.log_excel_file)
        except PermissionError:
            pass
        book.close()

    def add_progress_bar_to_status_bar(self):
        # Create a panel for embedding controls in the status bar
        self.progress_panel = wx.Panel(self.status_bar)
        self.progress_panel.SetBackgroundColour(self.status_bar.GetBackgroundColour())

        # Create a progress bar in determinate mode
        self.progress_bar = wx.Gauge(self.progress_panel, range=100, style=wx.GA_HORIZONTAL)
        sizer = wx.BoxSizer(wx.HORIZONTAL)
        sizer.Add(self.progress_bar, 1, wx.EXPAND | wx.ALL, 1)
        self.progress_panel.SetSizer(sizer)

        # Set the initial position and size of the panel
        self.update_progress_bar_position()

    def update_progress_bar_position(self):
        """Update the position and size of the progress panel based on the status bar field."""
        rect = self.status_bar.GetFieldRect(4)  # Field index 3 (last field)
        self.progress_panel.SetPosition((rect.x, rect.y))
        self.progress_panel.SetSize((rect.width, rect.height))
        self.progress_panel.Layout()

    def on_resize(self, event):
        """Handle window resizing and adjust the progress bar's size."""
        self.update_progress_bar_position()
        event.Skip()  # Ensure the default resize behavior happens

    @staticmethod
    def get_images_choices(choices_exo):
        result = []
        for choice in choices_exo:
            result.append(choice['image'])
        return result

    def display_exo(self):
        print("---->>display_exo")

        # print("*" * 50)
        # print("   ---->>Verification Start")
        # for ttt in self.all_exo[self.current_exo_name]:
        #     print(f"     ->{ttt}: {self.all_exo[self.current_exo_name][ttt]}")
        # print("   ---->>Verification End")

        if self.stage_type == "choices":
            self.valiny.Hide()
            self.ok_button.Hide()
            choices_config = self.get_images_choices(self.all_exo[self.current_exo_name]['choices'])
            choices_exo = self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['choices']
            if len(choices_config) > 0:
                self.change_bitmap_buttons(choices_config)
            elif len(choices_exo) > 0:
                self.change_bitmap_buttons(choices_exo)
            else:
                print(f"-->VER01")
                print(f"-1->{choices_config}")
                print(f"-2->{choices_exo}")
        elif self.stage_type == "entry":
            self.hide_bitmap_buttons()
            self.valiny.Show()
            print("----->TEST01")
            self.valiny.SetValue("")
            self.valiny.SetFocus()
            self.ok_button.Show()

        self.background_staticbitmap.Show()
        self.video_panel.Hide()
        self.home_panel.Layout()

        # self.player.player.stop()
        print("    -->>", "image1 : ", type(self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['image1']))
        print("    -->>", "image2 : ", type(self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['image2']))
        print("    -->>", "choices : ", self.all_exo[self.current_exo_name]['choices'])
        mp3 = self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['mp3']
        print("    -->>", "mp3 : ", mp3)
        print("    -->>", "answer : ", self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['answer'])
        print("    -->>", "text : ", self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['text'])

        self.load_image(self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['image1'])
        self.background_staticbitmap.SetCursor(wx.Cursor(wx.CURSOR_HAND))
        self.play_combined_mp3(mp3)

    def get_exo_index(self):
        print("--->>get_exo_index")
        ret = ""
        list1 = list(range(len(self.all_exo[self.current_exo_name]["exo"])))
        list2 = self.stage_index_done
        r = self.all_exo[self.current_exo_name]["rand"]
        nd = [x for x in list1 if x not in list2]
        print("     full: ", list1)
        print("     done: ", self.stage_index_done)
        print("     not done: ", nd)
        if 0 < len(nd) and len(self.stage_index_done) < self.stage_min:
            if r:
                ret = random.choice(nd)
            else:
                ret = nd[0]
        else:
            ret = None
        self.stage_current_index = ret
        print("     -->index =", ret)

    def get_exo_name(self):
        print("-" * 100)
        print("--->>get_exo_name")
        self.stage_current_index = None
        list1 = self.exo_list
        list2 = self.exo_done
        ret = ""

        try:
            self.progress_bar.SetValue(int(len(self.exo_done) / len(self.exo_list) * 100))
        except ZeroDivisionError:
            pass

        # print(f"     --VS-->>{self.exo_done} VS {self.exo_list}")

        # TODO: randomize the exo
        r = False
        nd = [x for x in list1 if x not in list2]
        if len(nd) > 0:
            if r:
                print(f"     -->exo Name: {nd[0]}")
                tmp_exo_name = nd[random.choice(list1)]
                ret = tmp_exo_name
                self.exo_done.append(ret)
            else:
                print(f"     -->exo Name: {nd[0]}")
                tmp_exo_name = nd[0]
                ret = nd[0]
                self.exo_done.append(ret)
            case_sensitive = ""
            if self.stage_case_sensitive:
                case_sensitive = ", Case_sst"
            self.SetStatusText(f"Exercise: {tmp_exo_name} (Level={self.stage_level}{case_sensitive})", 1)
        else:
            self.all_exo_completed = True
            print("Completed2")
            self.SetStatusText(f"Exercise: Completed", 1)
            self.display_exo_complete()
            ret = None
        self.current_exo_name = ret
        self.SetStatusText(f"Level: {len(self.exo_done)}/{len(self.exo_list)}", 3)

    def verify_correctness_all_exo(self):
        print("---->>verify_correctness_all_exo")
        new_all_exo = {}
        for exo in self.exo_list:
            # Verification
            # print("*"*50)
            # print(f"   ---->>Verification Start: {exo}")
            # # print(f"   ---->>Nb exo: {len(self.all_exo[exo]['exo'])}")
            # for ttt in self.all_exo[exo]:
            #     print(f"   ->{ttt}: {type(self.all_exo[exo][ttt])}")
            #     if ttt == "exo":
            #         print(f"        ---->>Nb exo: {len(self.all_exo[exo]['exo'])}")
            # print("   ---->>Verification End")
            try:
                if not len(self.all_exo[exo]["exo"]) == 0:
                    new_all_exo[exo] = self.all_exo[exo]
                    print(f"    -->Correct '{exo}'")
                else:
                    print(f"    -1->Removed '{exo}'")
                    print(f"    -1->Removed '{self.all_exo[exo]['exo']}'")

            except KeyError as e:
                print(f"    -2->Removed '{exo}' --> {e}")
        self.all_exo = new_all_exo
        self.exo_list = list(self.all_exo)

    @staticmethod
    def split_subject_answer(subject, answer):
        tmp = f"{subject}".split("|")
        if len(tmp) == 2:
            tmp_answer = f"{tmp[1]}".split("\t")
            tmp_answer_dic = {}
            for i in range(len(tmp_answer)):
                tmp_answer_dic[f"C{i + 1}"] = tmp_answer[i]
            if answer in tmp_answer_dic:
                return tmp[0], tmp_answer_dic[answer]
        return subject, answer

    def verify_answer(self, user_answer=""):
        print("--->>verify_answer")
        print(f"    --1->>{self.current_exo_name}")
        print(f"    --2->>{self.stage_current_index}")
        if user_answer == "":
            user_answer = self.valiny.GetValue()
        exo_answer = self.all_exo[self.current_exo_name]["exo"][self.stage_current_index]["answer"]
        print(f"    '{user_answer}' VS '{exo_answer}'")

        match = exo_answer.search(user_answer)
        if f"{user_answer}".strip() == "":
            # self.player.play_media(r"mp3/wrong.mp3")
            # time.sleep(1)
            pass
        elif match:
            self.valiny.SetValue("")
            self.stage_index_done.append(self.stage_current_index)
            print("    --->>MARINA")
            self.SetStatusText("MARINA !")
            self.valiny.SetValue("")
            self.Refresh()
            self.player.play_media(r"mp3/right.mp3")
            self.load_image(self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['image2'])
            # self.valiny.Hide()
            self.home_panel.Layout()
            # time.sleep(20)
            # self.valiny.Show()
            # self.valiny.Enable(True)
            subject = self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['text']
            subject, user_answer = self.split_subject_answer(subject, user_answer)
            self.update_log_file(["Marina", self.current_exo_name, self.stage_current_index, subject, user_answer])

            time.sleep(2)
            return True
        else:
            print("    --->>DISO")
            self.SetStatusText("DISO !")
            self.valiny.SetValue("")
            self.player.play_media(r"mp3/wrong.mp3")
            # time.sleep(20)
            if self.stage_min < self.stage_max:
                self.stage_min += 1
            subject = self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['text']
            subject, user_answer = self.split_subject_answer(subject, user_answer)
            self.update_log_file(["Diso", self.current_exo_name, self.stage_current_index, subject, user_answer])
            return False

    def get_level_config(self):
        print("---->>get_level_config")
        if self.current_exo_name in self.exo_list:
            self.stage_min = self.all_exo[self.current_exo_name]["min"]
            self.stage_max = self.all_exo[self.current_exo_name]["max"]
            self.stage_rand = self.all_exo[self.current_exo_name]["rand"]
            self.stage_level = self.all_exo[self.current_exo_name]["level"]
            self.stage_type = self.all_exo[self.current_exo_name]["type"]
            self.stage_case_sensitive = self.all_exo[self.current_exo_name]["case sensitive"]
            self.stage_library = self.all_exo[self.current_exo_name]["library"]
            self.stage_comment = self.all_exo[self.current_exo_name]["comment"]

            print(f"     ->stage_min: {self.stage_min}")
            print(f"     ->stage_max: {self.stage_max}")
            print(f"     ->stage_rand: {self.stage_rand}")
            print(f"     ->stage_level: {self.stage_level}")
            print(f"     ->stage_type: {self.stage_type}")
            print(f"     ->stage_case_sensitive: {self.stage_case_sensitive}")
            print(f"     ->library: {self.stage_library}")
            print(f"     ->stage_comment: {self.stage_comment}")

            directory = os.path.join('images', self.current_exo_name)
            course_video = ""
            files = [f for f in os.listdir(directory)]
            for file in files:
                if re.search("course.*\\.(mp4|avi)$", file, re.I):
                    course_video = os.path.join(directory, file)
                    break

            if os.path.isfile(course_video):
                self.background_staticbitmap.Hide()
                self.playing_video = True
                self.valiny.Hide()
                self.ok_button.Hide()
                self.video_panel.Show()
                self.player.play_media(course_video)
                self.home_panel.Layout()
            else:
                if self.current_exo_name is not None:
                    self.get_exo_index()
                    if self.stage_current_index is not None:
                        self.display_exo()

    def create_bitmap_buttons(self):
        for index in range(10):
            img = wx.Image(100, 100)
            bitmap = wx.Bitmap(img)
            bitmap_button = wx.BitmapButton(self.home_panel, -1, bitmap)
            bitmap_button.SetMinSize(bitmap.GetSize())
            bitmap_button.Bind(wx.EVT_BUTTON, self.on_bitmap_click)
            bitmap_button.Bind(wx.EVT_MOTION, self.on_bitmap_motion)
            bitmap_button.SetCursor(wx.Cursor(wx.CURSOR_HAND))
            bitmap_button.index = index
            self.bitmap_buttons.append(bitmap_button)
            self.bitmap_sizer.Add(bitmap_button, 0, wx.ALIGN_CENTER | wx.ALL, 5)
        self.choice_answer_available = True
        self.hide_bitmap_buttons()

    def on_home_panel_motion(self, event):
        self.dc.Clear()
        # self.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))

    def on_bitmap_motion(self, event):
        """Draws a rectangle around the hovered bitmap button."""
        bitmap_button = event.GetEventObject()
        # self.dc.SetPen(wx.Pen("red", 2))
        self.dc.SetBrush(wx.TRANSPARENT_BRUSH)
        x, y = bitmap_button.GetPosition()
        width, height = bitmap_button.GetSize()
        self.dc.DrawRectangle(x - 1, y - 1, width + 3, height + 3)

    def change_bitmap_buttons(self, new_img, max_height=150):
        print("---->>change_bitmap_buttons")
        self.hide_bitmap_buttons()
        for img in new_img:
            print(f"   -img->>{type(img)}")
        for index, img_data in enumerate(new_img):
            try:
                if isinstance(img_data, bytes):
                    width, height = 120, 80  # Default size, should be passed correctly
                    pil_img = Image.frombytes('RGB', (width, height), img_data)
                elif isinstance(img_data, Image.Image):
                    pil_img = img_data
                else:
                    full_path = os.path.join("images", img_data)
                    img = wx.Image(full_path, wx.BITMAP_TYPE_ANY)

                if isinstance(img_data, (bytes, Image.Image)):
                    stream = io.BytesIO()
                    pil_img.save(stream, format='PNG')
                    stream.seek(0)
                    img = wx.Image(stream, wx.BITMAP_TYPE_PNG)

                # Calculate scaled dimensions while maintaining aspect ratio
                width, height = img.GetSize()
                if height > max_height:
                    scale_factor = max_height / height
                    new_width = int(width * scale_factor)
                    img = img.Scale(new_width, max_height)

                bitmap = wx.Bitmap(img)
                self.bitmap_buttons[index].SetBitmapLabel(bitmap)
                self.bitmap_buttons[index].SetMinSize(bitmap.GetSize())
                self.bitmap_buttons[index].Show()
            except Exception as e:
                print(f"Error loading image: {e}")

    def hide_bitmap_buttons(self):
        """Hides all StaticBitmaps."""
        for static_bitmap in self.bitmap_buttons:
            static_bitmap.Hide()
        self.choice_answer_available = False

    def show_bitmap_buttons(self):
        """Shows all StaticBitmaps."""
        for static_bitmap in self.bitmap_buttons:
            static_bitmap.Show()
        self.choice_answer_available = True

    def OnSettings(self, event):
        # todo: fix issue, restarting while the video is playing
        dlg = Setting_DLG(self, title='Minatra Settings', style=wx.CLOSE_BOX | wx.CAPTION)
        dlg.CenterOnParent()
        res = dlg.ShowModal()
        if res == wx.ID_OK:
            json_path = os.path.join(os.getcwd(), "exo_schedule.json")
            with open(json_path, 'w') as outfile:
                json.dump(dlg.json_data, outfile, indent=3)
            exo_folder = dlg.TextCtrl_1.GetValue()
            if os.path.isdir(exo_folder):
                self.save_settings(exo_folder)

    def OnRestart(self, event):
        img_path = os.path.join("images", "orange_ice_mint.jpg")
        self.player.player.stop()
        # self.video_panel.Hide()
        if os.path.isfile(img_path):
            self.background_image = wx.Image(img_path, wx.BITMAP_TYPE_ANY)
        else:
            self.background_image = wx.Image(854, 480)
        if self.background_image.GetWidth() > 854:
            new_width = 854
            new_height = int(854 * self.background_image.GetHeight() / self.background_image.GetWidth())
            self.background_image = self.background_image.Scale(new_width, new_height, wx.IMAGE_QUALITY_HIGH)

        self.background_bitmap = wx.Bitmap(self.background_image)
        self.background_staticbitmap.Show()
        self.background_staticbitmap.SetBitmap(self.background_bitmap)
        self.home_panel.Refresh()

        self.ok_button.Show()
        self.valiny.Hide()
        self.ok_button.SetLabel("START")
        self.hide_bitmap_buttons()
        self.home_panel.Layout()
        self.ok_button.SetFocus()

    def OnExit(self, event):
        self.Close(True)

    def OnAbout(self, event):
        wx.MessageBox("Fianarana 1.0\n\nDeveloper: Njaka ANDRIAMAHENINA\nEmail: a6njaka@gmail.com", "About", wx.OK | wx.ICON_INFORMATION)

    def OnTool1(self, event):
        print("-->OnTool1 Clicked")
        txt_msg = ""
        for txt in self.exo_list:
            txt_msg += f"-{txt}\n"
        wx.MessageBox(txt_msg.strip())

    def OnTool2(self, event):
        print("-->OnTool2 Clicked")
        self.background_staticbitmap.Hide()
        self.player.play_media(r"D:\Videos\Facebook.mp4")
        # self.video_panel.Hide()
        # img_paths = [
        #     r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2\images\mividy_voankazo\source_images\annana-1000.png",
        #     r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2\images\mividy_voankazo\source_images\mangue-700.jpg",
        #     r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2\images\mividy_voankazo\source_images\Orange-600.jpg",
        #     r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2\images\mividy_voankazo\source_images\pastèque-2000.png",
        #     r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2\images\mividy_voankazo\source_images\poire-400.jpg",
        #     r"D:\Njaka_Project\Njaka_Dev_Itk\bin\Mianatra2\images\mividy_voankazo\source_images\pomme-500.jpg",
        # ]
        # self.change_bitmap_buttons(img_paths)
        self.ok_button.Hide()
        self.home_panel.Layout()

    def load_image(self, img_path):
        if isinstance(img_path, str) and os.path.isfile(img_path):
            new_image = wx.Image(img_path, wx.BITMAP_TYPE_ANY)
            if new_image.GetWidth() > 854:
                new_width = 854
                new_height = int(854 * new_image.GetHeight() / new_image.GetWidth())
                new_image = new_image.Scale(new_width, new_height, wx.IMAGE_QUALITY_HIGH)
            new_bitmap = wx.Bitmap(new_image)
            self.background_staticbitmap.SetBitmap(new_bitmap)
            self.home_panel.Refresh()

        # Bytes_Type
        elif isinstance(img_path, bytes):
            # if type(img_path) is bytes:
            # image_data = lib_addition_3ch_hor.get_image(0)
            image_data = img_path
            wx_image = wx.Image(854, 480)
            wx_image.SetData(image_data)
            wx_bitmap = wx.Bitmap(wx_image)
            self.background_staticbitmap.SetBitmap(wx_bitmap)
            self.home_panel.Refresh()
        elif isinstance(img_path, Image.Image):
            img_path = img_path.convert("RGB")
            width, height = img_path.size
            data = img_path.tobytes()
            new_image = wx.Image(width, height)
            new_image.SetData(data)
            # if new_image.GetWidth() > 854:
            #     new_width = 854
            #     new_height = int(854 * new_image.GetHeight() / new_image.GetWidth())
            #     new_image = new_image.Scale(new_width, new_height, wx.IMAGE_QUALITY_HIGH)
            new_bitmap = wx.Bitmap(new_image)
            self.background_staticbitmap.SetBitmap(new_bitmap)
            self.home_panel.Refresh()

        # self.SetStatusText("Background image changed.")
        self.home_panel.Layout()

    def load_all_exo(self):
        print("---->>load_all_exo")
        p1 = load_exo.ExoSchedule()
        self.all_exo = p1.all_exo
        self.exo_list = list(self.all_exo)
        self.all_exo_completed = False

    def on_ok_button(self, event):
        print(f"---->>OK_BUTTON")

        if self.ok_button.GetLabel() == "START":
            print("  -->>START")
            print(f"Exo Folder: {os.getcwd()}")
            self.ok_button.SetLabel("OK")
            self.exo_done = []
            self.load_all_exo()
            self.verify_correctness_all_exo()
            self.stage_index_done = []
            self.get_exo_name()
            self.get_level_config()
            # if self.current_exo_name is not None:
            #     self.get_exo_index()
            #     if self.stage_current_index is not None:
            #         self.display_exo()

        elif self.current_exo_name is None and self.ok_button.GetLabel() != "BRAVO !":
            self.display_exo_complete()
        elif self.ok_button.GetLabel() == "OK":
            if self.verify_answer():
                self.get_exo_index()
            if self.stage_current_index is not None:
                self.display_exo()
            else:
                self.stage_index_done = []
                self.get_exo_name()
                self.get_level_config()
                # if self.current_exo_name is not None:
                #     self.get_exo_index()
                #     if self.stage_current_index is not None:
                #         self.display_exo()
        self.SetStatusText(f"Stage {len(self.stage_index_done) + 1}/{self.stage_min} (Max : {self.stage_max})", 2)

    def display_exo_complete(self):
        # TODO: Display Bravo image
        self.ok_button.SetLabel("BRAVO !")
        img_path = os.path.join("images", "Bravo.jpg")
        self.player.play_media(r"mp3/bravo.mp3")
        self.load_image(img_path)
        self.valiny.Hide()
        self.ok_button.Show()
        self.hide_bitmap_buttons()
        self.home_panel.Layout()

    def on_background_click(self, event):
        all_mp3 = []
        try:
            files_exist = True
            all_mp3 = self.all_exo[self.current_exo_name]['exo'][self.stage_current_index]['mp3']
            if len(all_mp3) > 0:
                for mp3 in all_mp3:
                    if not os.path.isfile(mp3):
                        files_exist = False
                        break
                if files_exist:
                    self.play_combined_mp3(all_mp3)
        except:
            print(f"MP3 not correct! ---> {all_mp3}")

        # wx.MessageBox("Image Clicked", "Info", wx.OK | wx.ICON_INFORMATION)

    def on_enter_pressed(self, event):
        self.on_ok_button(event)
        # self.valiny.SetValue("Mety Tsara")  # Display the message when Enter is pressed

    def on_bitmap_click(self, event):
        clicked_bitmap = event.GetEventObject()
        index = clicked_bitmap.index

        # wx.MessageBox(f"Vous avez choisi l'image {index + 1}!", "Info", wx.OK | wx.ICON_INFORMATION)

        user_choice = f"C{index + 1}"
        if self.verify_answer(user_choice):
            self.get_exo_index()
        if self.stage_current_index is not None:
            self.display_exo()
        else:
            self.stage_index_done = []
            self.get_exo_name()
            self.get_level_config()
            # if self.current_exo_name is not None:
            #     self.get_exo_index()
            #     if self.stage_current_index is not None:
            #         self.display_exo()
        self.SetStatusText(f"Stage {len(self.stage_index_done) + 1}/{self.stage_min} (Max : {self.stage_max})", 2)

        clicked_bitmap.Refresh()
        self.home_panel.Layout()

    @staticmethod
    def play_combined_mp3(mp3_files):
        files_exist = True
        if len(mp3_files) > 0:
            for mp3 in mp3_files:
                if not os.path.isfile(mp3):
                    files_exist = False
                    break
        if files_exist and len(mp3_files) > 0:
            instance = vlc.Instance()
            player = instance.media_player_new()
            media_list = instance.media_list_new()

            for mp3_file in mp3_files:
                media = instance.media_new(mp3_file)
                media_list.add_media(media)

            list_player = instance.media_list_player_new()
            list_player.set_media_player(player)
            list_player.set_media_list(media_list)

            list_player.play()

            # Wait for playback to finish
            while list_player.get_state() != vlc.State.Ended:
                time.sleep(0.1)

            # Close the player after playback
            list_player.stop()
            list_player.release()
            player.release()
            instance.release()
        # else:
        #     print("MP3 not correct!")
        #     print(f"MP3 not correct! --2-> {mp3_files}")


class Setting_DLG(wx.Dialog):
    def __init__(self, *args, **kw):
        super(Setting_DLG, self).__init__(*args, **kw)
        self.SetSize((640, 460))
        self.list_all_exo = []
        self.list_group_exo = []
        self.exo_schedule = "exo_schedule.json"
        self.json_data = {}

        self.StaticText_cwd = wx.StaticText(self, wx.ID_ANY, "Exo folder:", wx.Point(32, 35), wx.Size(60, 15))
        self.TextCtrl_1 = wx.TextCtrl(self, 124, f"{os.getcwd()}", wx.Point(100, 32), wx.Size(380, 21), wx.TE_READONLY, wx.DefaultValidator, "ID_TEXTCTRL1")
        self.Browse_button = wx.Button(self, wx.ID_ANY, "Browse ...", wx.Point(500, 30), wx.Size(90, 25), 0, wx.DefaultValidator, "ID_BUTTON1")

        # Create controls
        self.Choice_group = wx.Choice(self, wx.ID_ANY, pos=(32, 70), size=(256, 21))
        self.CheckBox_enable = wx.CheckBox(self, wx.ID_ANY, "Enable", pos=(328, 72))
        self.StaticText2 = wx.StaticText(self, wx.ID_ANY, "Max exo:", pos=(400, 72))
        self.SpinCtrl_max_exo = wx.SpinCtrl(self, wx.ID_ANY, "0", pos=(460, 68), size=(50, 21), style=wx.SP_ARROW_KEYS, min=0, max=100)  # Important to set min and max here
        self.SpinCtrl_max_exo.SetValue(1)  # Set the initial value as an integer

        self.Choice_group.Append("Groupe_1")
        self.Choice_group.Append("Groupe_2")
        self.Choice_group.Append("Groupe_3")
        self.Choice_group.Append("Groupe_4")
        self.Choice_group.Append("Groupe_5")
        self.Choice_group.Append("Groupe_6")
        self.Choice_group.SetSelection(0)  # Select the first item

        self.StaticBox1 = wx.StaticBox(self, wx.ID_ANY, "Details", pos=(24, 100), size=(580, 250))

        self.Button_Exo_Add = wx.Button(self, wx.ID_ANY, ">>", pos=(232, 170))
        self.Button_Exo_Remove = wx.Button(self, wx.ID_ANY, "<<", pos=(232, 210))

        self.ListBox_available_exo = wx.ListBox(self, wx.ID_ANY, pos=(32, 120), size=(184, 220))
        self.ListBox_group_exo = wx.ListBox(self, wx.ID_ANY, pos=(328, 120), size=(176, 220))

        self.CheckBox_monday = wx.CheckBox(self, wx.ID_ANY, "Monday", pos=(520, 121))
        self.CheckBox_Tuesday = wx.CheckBox(self, wx.ID_ANY, "Tuesday", pos=(520, 153))
        self.CheckBox_wednesday = wx.CheckBox(self, wx.ID_ANY, "Wednesday", pos=(520, 185))
        self.CheckBox_thursday = wx.CheckBox(self, wx.ID_ANY, "Thursday", pos=(520, 217))
        self.CheckBox_friday = wx.CheckBox(self, wx.ID_ANY, "Friday", pos=(520, 249))
        self.CheckBox_saturday = wx.CheckBox(self, wx.ID_ANY, "Saturday", pos=(520, 281))
        self.CheckBox_sunday = wx.CheckBox(self, wx.ID_ANY, "Sunday", pos=(520, 313))

        self.text_exo_info = wx.StaticText(self, wx.ID_ANY, "", wx.Point(32, 355), wx.Size(560, 15))
        self.number_total_exo = wx.StaticText(self, wx.ID_ANY, "Total exo: 0", wx.Point(32, 375), wx.Size(120, 15))

        self.Button_OK = wx.Button(self, wx.ID_OK, "OK", pos=(328, 380), size=(104, 23))
        self.Button_Cancel = wx.Button(self, wx.ID_CANCEL, "Cancel", pos=(448, 380), size=(104, 23))

        # Event Handlers
        self.Bind(wx.EVT_BUTTON, self.on_button_add, self.Button_Exo_Add)
        self.Bind(wx.EVT_BUTTON, self.on_button_remove, self.Button_Exo_Remove)
        self.Bind(wx.EVT_BUTTON, self.on_browse_button, self.Browse_button)
        self.Bind(wx.EVT_LISTBOX, self.listbox_available_exo_click, self.ListBox_available_exo)
        self.Bind(wx.EVT_LISTBOX, self.listbox_group_exo_click, self.ListBox_group_exo)

        self.StaticText_cwd.Bind(wx.EVT_LEFT_DCLICK, self.on_cwd_double_click)
        self.Choice_group.Bind(wx.EVT_CHOICE, self.on_choice_group_changed)
        self.SpinCtrl_max_exo.Bind(wx.EVT_SPINCTRL, self.update_json)
        self.CheckBox_monday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_Tuesday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_wednesday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_thursday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_friday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_saturday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_sunday.Bind(wx.EVT_CHECKBOX, self.update_json)
        self.CheckBox_enable.Bind(wx.EVT_CHECKBOX, self.on_checkbox_enable)

        self.update_list_exo()
        self.read_json_exo_schedule()
        self.update_total_exo_txt()
        self.on_choice_group_changed(None)
        self.Choice_group.SetFocus()

    def on_checkbox_enable(self, event):
        self.update_json(event)
        self.Choice_group.SetFocus()


    def update_json(self, event):
        group_tmp = {
            "activate": self.CheckBox_enable.GetValue(),
            "exo_group": [self.ListBox_group_exo.GetString(i) for i in range(self.ListBox_group_exo.GetCount())],
            "exo_number": self.SpinCtrl_max_exo.GetValue(),
            "exo_group_name": self.Choice_group.GetStringSelection(),
            "exo_weekdays": []
        }
        group_index = self.Choice_group.GetSelection()

        if self.CheckBox_monday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_monday.GetLabel())
        if self.CheckBox_Tuesday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_Tuesday.GetLabel())
        if self.CheckBox_wednesday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_wednesday.GetLabel())
        if self.CheckBox_thursday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_thursday.GetLabel())
        if self.CheckBox_friday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_friday.GetLabel())
        if self.CheckBox_saturday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_saturday.GetLabel())
        if self.CheckBox_sunday.GetValue():
            group_tmp["exo_weekdays"].append(self.CheckBox_sunday.GetLabel())

        self.json_data["Itokiana"][group_index] = group_tmp
        self.update_total_exo_txt()


    def update_total_exo_txt(self):
        total_exo = 0
        for exo in self.json_data["Itokiana"]:
            try:
                if exo["activate"]:
                    min_val = min(exo["exo_number"], len(exo["exo_group"]))
                    total_exo += min_val
            except Exception as e:
                print(f"Error total exo = {e}")
        self.number_total_exo.SetLabel(f"Total exo: {total_exo}")


    def on_choice_group_changed(self, event):
        index = self.Choice_group.GetSelection()
        self.read_and_update_group_exo(index)
        self.update_list_exo()

    def on_browse_button(self, event):
        dirdialog1 = wx.DirDialog(self, "Select Exo Folder", wx.EmptyString,
                                  wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST, wx.DefaultPosition, wx.DefaultSize,
                                  "wxDirDialog")
        dlg = dirdialog1.ShowModal()
        if dlg == wx.ID_OK:
            new_dir = dirdialog1.GetPath()
            self.TextCtrl_1.SetValue(new_dir)
            os.chdir(new_dir)

    def listbox_available_exo_click(self, event):
        self.ListBox_group_exo.Deselect(self.ListBox_group_exo.GetSelection())
        info = self.get_exo_info(self.ListBox_available_exo.GetStringSelection())
        self.text_exo_info.SetLabel(info)

    def listbox_group_exo_click(self, event):
        self.ListBox_available_exo.Deselect(self.ListBox_available_exo.GetSelection())
        info = self.get_exo_info(self.ListBox_group_exo.GetStringSelection())
        self.text_exo_info.SetLabel(info)

    def get_exo_info(self, exo):
        exo_path = os.path.join(self.TextCtrl_1.GetValue(), "images", exo)
        ret = []
        if os.path.isdir(exo_path):
            json_config = os.path.join(f"{exo_path}", "config.json")
            if os.path.isfile(json_config):
                with open(json_config) as json_file:
                    data = json.load(json_file)
                    if "min" in data:
                        ret.append(f"Min= {data['min']}")
                    if "max" in data:
                        ret.append(f"Max= {data['max']}")
                    if "level" in data:
                        ret.append(f"Level= {data['level']}")
                    if "creation date" in data:
                        ret.append(f"Date= {data['creation date']}")
        info = ", ".join(map(str, ret))
        exo = f"{exo}".upper()
        return f"{exo}: {info}"

    @staticmethod
    def sort_list_box(listbox):
        items = [listbox.GetString(i) for i in range(listbox.GetCount())]
        items.sort()  # In place sort
        listbox.Set(items)

    def read_json_exo_schedule(self):
        group_tmp = {
            "activate": False,
            "exo_group": [],
            "exo_number": 0,
            "exo_group_name": "",
            "exo_weekdays": []
        }
        with open(self.exo_schedule) as json_file:
            self.json_data = json.load(json_file)
            if "Itokiana" in self.json_data:
                n = len(self.json_data["Itokiana"])
                max_group = 6
                for i in range(max_group + 1):
                    if i > n:
                        self.json_data["Itokiana"].append(group_tmp)

    def read_and_update_group_exo(self, group_index):
        self.ListBox_group_exo.Clear()
        self.list_group_exo = []
        user = "Itokiana"
        if len(self.json_data[user]) > group_index:
            for exo in self.json_data[user][group_index]["exo_group"]:
                self.ListBox_group_exo.Append(exo)
                self.list_group_exo.append(exo)
            try:
                self.SpinCtrl_max_exo.SetValue(int(self.json_data[user][group_index]["exo_number"]))
            except Exception as e:
                print(f"Error exo_number: {e}")
            try:
                self.CheckBox_enable.SetValue(int(self.json_data[user][group_index]["activate"]))
            except Exception as e:
                print(f"Error activate: {e}")
            try:
                days = self.json_data[user][group_index]["exo_weekdays"]
                if "Monday" in days:
                    self.CheckBox_monday.SetValue(True)
                else:
                    self.CheckBox_monday.SetValue(False)
                if "Tuesday" in days:
                    self.CheckBox_Tuesday.SetValue(True)
                else:
                    self.CheckBox_Tuesday.SetValue(False)
                if "Wednesday" in days:
                    self.CheckBox_wednesday.SetValue(True)
                else:
                    self.CheckBox_wednesday.SetValue(False)
                if "Thursday" in days:
                    self.CheckBox_thursday.SetValue(True)
                else:
                    self.CheckBox_thursday.SetValue(False)
                if "Friday" in days:
                    self.CheckBox_friday.SetValue(True)
                else:
                    self.CheckBox_friday.SetValue(False)

                if "Saturday" in days:
                    self.CheckBox_saturday.SetValue(True)
                else:
                    self.CheckBox_saturday.SetValue(False)

                if "Sunday" in days:
                    self.CheckBox_sunday.SetValue(True)
                else:
                    self.CheckBox_sunday.SetValue(False)

            except Exception as e:
                print(f"Error exo_weekdays: {e}")

        else:
            self.SpinCtrl_max_exo.SetValue(0)
            self.CheckBox_enable.SetValue(False)
            self.CheckBox_monday.SetValue(False)
            self.CheckBox_Tuesday.SetValue(False)
            self.CheckBox_wednesday.SetValue(False)
            self.CheckBox_thursday.SetValue(False)
            self.CheckBox_friday.SetValue(False)
            self.CheckBox_saturday.SetValue(False)
            self.CheckBox_sunday.SetValue(False)
        self.sort_list_box(self.ListBox_available_exo)
        self.sort_list_box(self.ListBox_group_exo)

    @staticmethod
    def on_cwd_double_click(event):
        cwd = os.getcwd()
        try:
            # Windows: Use subprocess to open File Explorer
            subprocess.Popen(["explorer", cwd])
        except Exception as e:
            wx.MessageBox(f"Error opening directory: {e}", "Error", wx.OK | wx.ICON_ERROR)

        event.Skip()  # Important: Allow other events to process

    def on_button_add(self, event):
        selected_item = self.ListBox_available_exo.GetSelection()
        if selected_item != wx.NOT_FOUND:
            item_text = self.ListBox_available_exo.GetString(selected_item)
            self.ListBox_group_exo.Append(item_text)
            self.ListBox_available_exo.Delete(selected_item)
        self.sort_list_box(self.ListBox_available_exo)
        self.sort_list_box(self.ListBox_group_exo)
        self.update_json(None)

    def on_button_remove(self, event):
        selected_item = self.ListBox_group_exo.GetSelection()
        if selected_item != wx.NOT_FOUND:
            item_text = self.ListBox_group_exo.GetString(selected_item)
            self.ListBox_available_exo.Append(item_text)
            self.ListBox_group_exo.Delete(selected_item)
        self.sort_list_box(self.ListBox_available_exo)
        self.sort_list_box(self.ListBox_group_exo)
        self.update_json(None)

    def update_list_exo(self):
        self.ListBox_available_exo.Clear()
        all_exo_path = os.path.join(os.getcwd(), "images")
        if not os.path.exists(all_exo_path):
            return None

        for exo in os.listdir(all_exo_path):
            # item_path = os.path.join(exo_path, item)
            exo_path = os.path.join(all_exo_path, exo)
            if os.path.isdir(exo_path):
                self.list_all_exo.append(exo)
                if exo not in self.list_group_exo:
                    self.ListBox_available_exo.Append(exo)


if __name__ == "__main__":
    app = wx.App()
    frame = MyFrame(None, "FIANARANA 1.0")
    app.MainLoop()
