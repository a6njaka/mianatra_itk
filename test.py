
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


log_excel_file = "log_exo_itokiana.xlsx"
# book = load_workbook(log_excel_file)
book = Workbook()

sheet = book.active

# ------------------Insert data -------------------------
sheet.append(("DOB", "LAST NAME", "FIRST NAME"))
sheet['A2'] = "06/01/1980"
sheet['B2'] = "ANDRIAMAHENINA"
sheet.cell(row=2, column=3).value = "Njaka Namelantsoa"

sheet.append(("04/07/1983", "TAHINAMARINELA", "Lanjasoa Manampihanitra"))
sheet.append(("24/04/2009", "ANDRIAMAHENINA", "Anjarasoa Itokiana"))
sheet.append(("12/10/2012", "ANDRIAMAHENINA", "Miotisoa Ifaliana"))
sheet.append(("24/12/2016", "ANDRIAMAHENINA", "Iarovana Nomenasoa"))
sheet.title = "Family"
sheet2 = book.create_sheet("Data")
sheet2.append(("06/01/1980", "ANDRIAMAHENINA", "Njaka Namelantsoa"))

# -------------------- Layout -------------------------
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
column_width = {"A": 15, "B": 25, "C": 30}
for row in range(sheet.max_row):
    for col in range(len(column_width)):
        sheet.cell(row + 1, col + 1).border = thin_border

for col in column_width:
    sheet.column_dimensions[col].width = column_width[col]
    sheet[f"{col}1"].font = Font(bold=True)
    sheet[f"{col}1"].fill = PatternFill(start_color="00A9E6", fill_type="solid")
    sheet[f"{col}1"].border = thin_border

# ------------------------- Save ----------------------
book.save(log_excel_file)
book.close()

# pyinstaller -w -F Mianatra2.py